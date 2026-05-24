"""Multi-worksheet Excel ingestion — dashboard + raw data pattern."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.analysis import load_taxonomy
from app.services.ingestion import (
    infer_tabular_schema,
    parse_spend_file_with_report,
    score_workbook_sheets,
    workbook_schema_hints,
)


def _build_dashboard_raw_workbook(path: Path) -> None:
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash["A1"] = "GCC OpEx Dashboard"
    dash["A2"] = "Total Spend KPI"
    dash["B2"] = 50000000
    dash["A3"] = "FTE Count"
    dash["B3"] = 1200

    raw = wb.create_sheet("Raw Data")
    raw.append(["Supplier", "Description", "Amount", "Category"])
    raw.append(["Infosys Ltd", "Application support", 2500000, "IT"])
    raw.append(["WeWork", "Facility lease", 800000, "FACILITIES"])
    raw.append(["NASSCOM Member", "Industry membership", 50000, "PROF_SVCS"])
    raw.append(["TCS", "BPO services", 1200000, "OUTSOURCED"])
    raw.append(["AWS India", "Cloud hosting", 950000, "IT"])

    wb.save(path)


@pytest.fixture
def dashboard_raw_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "gcc_dashboard_raw.xlsx"
    _build_dashboard_raw_workbook(p)
    return p


def test_score_workbook_prefers_raw_data_sheet(dashboard_raw_xlsx: Path) -> None:
    ranked = score_workbook_sheets(dashboard_raw_xlsx)
    assert len(ranked) >= 2
    assert ranked[0].sheet_name == "Raw Data"
    assert ranked[0].has_amount_col
    assert ranked[0].score > ranked[1].score


def test_workbook_schema_hints_select_raw_data(dashboard_raw_xlsx: Path) -> None:
    hints = workbook_schema_hints(dashboard_raw_xlsx)
    assert hints["sheet_count"] == 2
    assert hints["selected_sheet"] == "Raw Data"


def test_parse_spend_file_ingests_raw_data_not_dashboard(dashboard_raw_xlsx: Path) -> None:
    taxonomy = load_taxonomy()
    lines, report = parse_spend_file_with_report(dashboard_raw_xlsx, taxonomy)
    assert len(lines) >= 4
    assert report["sheets_ingested"]
    assert report["sheets_ingested"][0]["sheet"] == "Raw Data"
    assert report["sheets_ingested"][0]["rows"] == len(lines)
    skipped_names = {s["sheet"] for s in report.get("sheets_skipped", [])}
    assert "Dashboard" in skipped_names
    suppliers = {ln.supplier for ln in lines}
    assert "Infosys Ltd" in suppliers or "AWS India" in suppliers


def test_spend_category_not_used_as_amount_column(tmp_path: Path) -> None:
    """Regression: Spend_Category must not shadow Amount_USD (GCC-style ledgers)."""
    path = tmp_path / "gcc_ledger.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Data"
    ws.append(
        ["Invoice_ID", "Date", "Vendor", "Spend_Category", "Department", "Region", "Amount_USD", "Status"]
    )
    ws.append(["INV-1", "2023-11-24", "HP", "IT Hardware", "IT Operations", "India - Bangalore", 4348.91, "Paid"])
    ws.append(["INV-2", "2023-10-07", "Microsoft", "Software & Cloud", "IT Operations", "India - Gurugram", 14106.63, "Paid"])
    wb.save(path)

    taxonomy = load_taxonomy()
    lines, report = parse_spend_file_with_report(path, taxonomy)
    assert len(lines) == 2
    assert lines[0].amount == pytest.approx(4348.91)
    assert lines[1].amount == pytest.approx(14106.63)
    assert lines[0].currency == "USD"
    assert lines[0].description == "IT Hardware"
    schema = infer_tabular_schema(path)
    assert schema["semantic_map"]["amount"] == "Amount_USD"
    assert report["sheets_ingested"][0]["rows"] == 2


def test_ledger_manifest_strategy(dashboard_raw_xlsx: Path) -> None:
    from app.skills.model_contextualizer import build_workbook_manifest

    manifest, _meta = build_workbook_manifest(
        dashboard_raw_xlsx,
        user_message="gcc spend upload",
        session_meta={"industry": "gcc_capability_centers"},
    )
    dumped = manifest.model_dump()
    roles = {n["role"] for n in dumped.get("sheet_graph", [])}
    assert "transaction_ledger" in roles or dumped.get("ingestion_strategy") == "ledger_standard"
    taxonomy = load_taxonomy()
    lines, report = parse_spend_file_with_report(
        dashboard_raw_xlsx, taxonomy, workbook_manifest=dumped
    )
    assert len(lines) >= 4
    assert report["sheets_ingested"][0]["sheet"] == "Raw Data"
