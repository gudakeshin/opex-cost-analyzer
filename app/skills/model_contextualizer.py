from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from app.config import ANTHROPIC_ENABLED
from app.skills.contracts import WorkbookManifest, validate_workbook_manifest

_ROLE_KEYWORDS: Dict[str, tuple[str, ...]] = {
    "assumptions": ("assumption", "input", "driver", "parameter"),
    "timeseries": ("plan", "opex", "build", "forecast", "projection", "budget"),
    "summary": ("summary", "p&l", "pnl", "ebitda", "results", "output", "dashboard"),
    "scenarios": ("scenario", "base", "bull", "bear", "upside", "downside"),
    "sensitivity": ("sensitivity", "what-if", "what if"),
    "cover": ("cover", "readme", "intro"),
    "helper": ("lookup", "helper", "mapping", "reference"),
}

_SCENARIO_LABELS = (
    "base",
    "base case",
    "plan",
    "budget",
    "bull",
    "upside",
    "optimistic",
    "best case",
    "bear",
    "downside",
    "pessimistic",
    "worst case",
    "conservative",
    "accelerated",
    "stretch",
)

_SUMMARY_METRICS = (
    "total opex",
    "total cost",
    "ebitda",
    "operating profit",
    "net savings",
    "cost per fte",
    "cost as % of revenue",
    "irr",
    "npv",
    "payback",
)

_DRIVER_PATTERNS: tuple[tuple[str, str], ...] = (
    ("headcount", "headcount_growth_pct"),
    ("fte growth", "headcount_growth_pct"),
    ("revenue growth", "revenue_growth_pct"),
    ("cost escalation", "price_escalation_pct"),
    ("inflation", "price_escalation_pct"),
    ("wacc", "discount_rate"),
    ("discount rate", "discount_rate"),
    ("execution rate", "execution_rate_pct"),
    ("realisation", "execution_rate_pct"),
    ("tax rate", "effective_tax_rate"),
)


def compute_file_fingerprint(files: List[Dict[str, Any]]) -> str:
    parts: List[str] = []
    for item in files:
        path = Path(str(item.get("path") or ""))
        if not path.exists():
            continue
        stat = path.stat()
        parts.append(f"{path.name}:{stat.st_size}:{int(stat.st_mtime)}")
    joined = "|".join(sorted(parts))
    return hashlib.sha256(joined.encode("utf-8")).hexdigest() if joined else ""


def should_run_model_contextualizer(files: List[Dict[str, Any]], user_message: str) -> bool:
    lowered_msg = (user_message or "").lower()
    if any(token in lowered_msg for token in ("planning model", "scenario model", "budget model", "forecast model", "multi-sheet")):
        return True
    excel_files = [f for f in files if Path(str(f.get("path", ""))).suffix.lower() in (".xlsx", ".xls")]
    if not excel_files:
        return False
    for f in excel_files:
        schema = f.get("schema", {})
        workbook = schema.get("workbook", {}) if isinstance(schema, dict) else {}
        if int(workbook.get("sheet_count", 0)) > 1:
            return True
        if float(workbook.get("planning_signal_confidence", 0.0) or 0.0) >= 0.5:
            return True
    return False


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_period(text: str) -> bool:
    t = (text or "").strip().lower()
    if not t:
        return False
    patterns = (
        r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\s\-_/]?\d{2,4}$",
        r"^fy\d{2,4}$",
        r"^q[1-4][\s\-_/]?\d{2,4}$",
        r"^\d{4}[\-_/]\d{1,2}$",
        r"^\d{4}$",
    )
    return any(re.match(p, t) for p in patterns)


def _period_grain(periods: List[str]) -> str:
    lowered = [p.lower() for p in periods if p]
    if not lowered:
        return "unknown"
    if all(p.startswith("q") for p in lowered):
        return "quarterly"
    if all(p.startswith("fy") or re.fullmatch(r"\d{4}", p) for p in lowered):
        return "annual"
    return "monthly"


def _extract_formula_refs(ws: Any, max_refs: int = 10) -> List[str]:
    refs: List[str] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 120), min_col=1, max_col=min(ws.max_column, 20)):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("=") and "!" in value:
                refs.append(value[:120])
                if len(refs) >= max_refs:
                    return refs
    return refs


def extract_structural_summary(
    workbook_path: Path,
    user_message: str,
    session_meta: Dict[str, Any],
) -> Dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    sheets: List[Dict[str, Any]] = []
    workbook_periods: List[str] = []
    for ws in wb.worksheets:
        first_rows: List[List[str]] = []
        for r in range(1, min(ws.max_row, 3) + 1):
            row_vals = [_clean_cell(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 10) + 1)]
            first_rows.append(row_vals)
        first_col_values = [
            _clean_cell(ws.cell(row=r, column=1).value)
            for r in range(1, min(ws.max_row, 20) + 1)
            if _clean_cell(ws.cell(row=r, column=1).value)
        ]
        for row_vals in first_rows:
            workbook_periods.extend([v for v in row_vals if _looks_like_period(v)])
        sheets.append(
            {
                "sheet_name": ws.title,
                "row_count": int(ws.max_row or 0),
                "col_count": int(ws.max_column or 0),
                "first_rows": first_rows,
                "first_col_values": first_col_values,
                "formula_refs": _extract_formula_refs(ws),
            }
        )
    named_ranges = [str(name) for name in wb.defined_names.keys()][:20]
    return {
        "workbook_name": workbook_path.name,
        "sheet_order": [s["sheet_name"] for s in sheets],
        "sheets": sheets,
        "named_ranges": named_ranges,
        "formula_reference_sample": [f for s in sheets for f in s.get("formula_refs", [])][:10],
        "user_message": user_message,
        "session_meta": {
            "industry": session_meta.get("industry"),
            "annual_revenue": session_meta.get("annual_revenue"),
            "currency": session_meta.get("currency"),
        },
        "period_samples": workbook_periods[:24],
    }


def _classify_sheet_role(sheet: Dict[str, Any]) -> Tuple[str, float]:
    name = str(sheet.get("sheet_name", "")).lower()
    first_col = " ".join(str(x).lower() for x in sheet.get("first_col_values", [])[:20])
    first_rows = " ".join(str(x).lower() for row in sheet.get("first_rows", []) for x in row[:10])
    haystack = f"{name} {first_col} {first_rows}"

    score_by_role: Dict[str, float] = {role: 0.0 for role in _ROLE_KEYWORDS}
    for role, words in _ROLE_KEYWORDS.items():
        for w in words:
            if w in haystack:
                score_by_role[role] += 0.25

    period_hits = sum(1 for token in sheet.get("first_rows", [[], []])[0] if _looks_like_period(token))
    if period_hits < 3 and len(sheet.get("first_rows", [])) > 1:
        period_hits = sum(1 for token in sheet.get("first_rows", [[], []])[1] if _looks_like_period(token))
    if period_hits >= 3:
        score_by_role["timeseries"] += 0.55

    if any(lbl in haystack for lbl in _SUMMARY_METRICS):
        score_by_role["summary"] += 0.5

    role = max(score_by_role.items(), key=lambda kv: kv[1])[0]
    confidence = max(score_by_role.values())
    if confidence < 0.35:
        return "unknown", 0.25
    return role, min(0.95, confidence)


def _detect_period_axis(sheet: Dict[str, Any]) -> Dict[str, Any] | None:
    first_rows = sheet.get("first_rows", []) or []
    for idx, row in enumerate(first_rows[:2], start=1):
        period_positions = [i + 1 for i, val in enumerate(row) if _looks_like_period(str(val))]
        if len(period_positions) >= 3:
            periods = [str(row[i - 1]) for i in period_positions]
            return {
                "orientation": "column",
                "first_period_col": period_positions[0],
                "last_period_col": period_positions[-1],
                "periods": periods,
                "header_row": idx,
            }
    return None


def _map_driver(label: str) -> str | None:
    low = label.lower()
    for pattern, target in _DRIVER_PATTERNS:
        if pattern in low:
            return target
    return None


def interpret_structure_heuristic(summary: Dict[str, Any]) -> Dict[str, Any]:
    sheet_graph: List[Dict[str, Any]] = []
    scenarios: List[Dict[str, Any]] = []
    key_drivers: List[Dict[str, Any]] = []
    output_metrics: List[Dict[str, Any]] = []
    coverage: List[str] = []
    periods: List[str] = []
    role_confidences: List[float] = []

    for sheet in summary.get("sheets", []):
        role, confidence = _classify_sheet_role(sheet)
        role_confidences.append(confidence)
        axis = _detect_period_axis(sheet) if role in {"timeseries", "scenarios"} else None
        if axis:
            periods.extend(axis.get("periods", []))
        row_labels = [str(x) for x in sheet.get("first_col_values", [])[:20] if x]
        for label in row_labels:
            mapped = _map_driver(label)
            if role == "assumptions" and mapped and len(key_drivers) < 20:
                key_drivers.append(
                    {
                        "variable_name": re.sub(r"\W+", "_", label.strip().lower()).strip("_"),
                        "source_sheet": sheet.get("sheet_name"),
                        "cell_ref": None,
                        "current_value": None,
                        "unit": None,
                        "maps_to_sensitivity_driver": mapped,
                    }
                )
            if role in {"summary", "scenarios"} and any(metric in label.lower() for metric in _SUMMARY_METRICS) and len(output_metrics) < 20:
                output_metrics.append(
                    {
                        "metric_name": re.sub(r"\W+", "_", label.strip().lower()).strip("_"),
                        "source_sheet": sheet.get("sheet_name"),
                        "cell_ref": None,
                        "value": None,
                        "currency": summary.get("session_meta", {}).get("currency"),
                        "scenario_label": None,
                        "pre_populated": False,
                    }
                )
        for lbl in row_labels:
            low = lbl.lower()
            if any(k in low for k in ("it", "technology", "facilit", "hr", "professional", "software", "cloud")):
                coverage.append(lbl)

        if role == "scenarios":
            labels = [x for x in row_labels if any(tok in x.lower() for tok in _SCENARIO_LABELS)]
            for i, label in enumerate(labels[:6]):
                scenarios.append(
                    {
                        "scenario_id": re.sub(r"\W+", "_", label.lower()).strip("_") or f"scenario_{i+1}",
                        "label": label,
                        "source_sheet": sheet.get("sheet_name"),
                        "column_index": None,
                        "maps_to_sensitivity_scenario": None,
                    }
                )

        sheet_graph.append(
            {
                "sheet_name": sheet.get("sheet_name", ""),
                "role": role,
                "feeds_into": [],
                "depends_on": [],
                "period_axis": axis,
                "driver_variables": [],
                "output_metrics": [],
                "categories_detected": [],
                "row_count": int(sheet.get("row_count", 0) or 0),
                "data_density": "sparse" if int(sheet.get("row_count", 0) or 0) < 40 else "dense" if int(sheet.get("row_count", 0) or 0) > 120 else "medium",
            }
        )

    roles = {node["role"] for node in sheet_graph}
    if "timeseries" in roles and "scenarios" in roles:
        ingestion_strategy = "hybrid"
    elif "timeseries" in roles and "assumptions" in roles:
        ingestion_strategy = "hybrid"
    elif "timeseries" in roles:
        ingestion_strategy = "timeseries_flatten"
    elif "scenarios" in roles:
        ingestion_strategy = "scenario_pivot"
    elif "assumptions" in roles:
        ingestion_strategy = "assumptions_extract"
    else:
        ingestion_strategy = "standard"

    model_type = "unknown"
    if "scenarios" in roles:
        model_type = "scenario"
    elif "timeseries" in roles and "assumptions" in roles:
        model_type = "planning"
    elif "timeseries" in roles:
        model_type = "forecast"

    unique_periods: List[str] = []
    for p in periods:
        if p and p not in unique_periods:
            unique_periods.append(p)
    planning_horizon = {
        "start_period": unique_periods[0] if unique_periods else None,
        "end_period": unique_periods[-1] if unique_periods else None,
        "period_grain": _period_grain(unique_periods),
        "total_periods": len(unique_periods),
    }

    confidence = sum(role_confidences) / len(role_confidences) if role_confidences else 0.0
    notes: List[str] = []
    if any(node["role"] == "unknown" for node in sheet_graph):
        notes.append("One or more sheets could not be confidently classified.")
    if not unique_periods:
        notes.append("No strong period axis detected from sampled header rows.")

    return {
        "manifest_version": "1.0",
        "workbook_name": summary.get("workbook_name", ""),
        "model_type": model_type,
        "planning_horizon": planning_horizon,
        "scenarios": scenarios,
        "sheet_graph": sheet_graph,
        "key_driver_variables": key_drivers[:20],
        "output_metrics": output_metrics[:20],
        "spend_category_coverage": sorted(set(coverage))[:25],
        "ingestion_strategy": ingestion_strategy,
        "ingestion_notes": " ".join(notes).strip(),
        "confidence": max(0.0, min(1.0, confidence)),
    }


def _interpret_structure_with_llm(summary: Dict[str, Any]) -> Tuple[Dict[str, Any] | None, str | None]:
    if not ANTHROPIC_ENABLED:
        return None, "provider_disabled"
    try:
        from app.opar.claude_client import interpret_workbook_structure_claude_with_meta
    except Exception:
        return None, "provider_unavailable"
    return interpret_workbook_structure_claude_with_meta(summary)


def build_workbook_manifest(
    workbook_path: Path,
    user_message: str,
    session_meta: Dict[str, Any],
    force_llm: bool = False,
) -> Tuple[WorkbookManifest, Dict[str, Any]]:
    summary = extract_structural_summary(workbook_path, user_message=user_message, session_meta=session_meta)
    heuristic = interpret_structure_heuristic(summary)
    llm_used = False
    fallback_reason = None

    raw_manifest = heuristic
    if force_llm or float(heuristic.get("confidence", 0.0)) < 0.70:
        llm_out, llm_reason = _interpret_structure_with_llm(summary)
        if llm_out:
            raw_manifest = llm_out
            llm_used = True
        else:
            fallback_reason = llm_reason or "llm_failed"

    try:
        parsed = validate_workbook_manifest(raw_manifest)
    except Exception:
        parsed = validate_workbook_manifest(heuristic)
        fallback_reason = fallback_reason or "validation_failed"
        llm_used = False

    meta = {
        "llm_used": llm_used,
        "fallback_reason": fallback_reason,
        "heuristic_confidence": float(heuristic.get("confidence", 0.0) or 0.0),
        "structural_summary": summary,
    }
    return parsed, meta

