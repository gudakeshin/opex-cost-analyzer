"""
PMO Toolkit Export — xlsx + RACI + milestone calendar + variance/FTC report.

Produces:
  1. Initiative tracker sheet (xlsx)
  2. RACI matrix sheet
  3. Milestone calendar sheet
  4. Forecast-to-Complete (FTC) variance report (JSON + xlsx sheet)
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, cast

from app.config import OUTPUT_DIR

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False


_RACI_ROLES = ["CFO", "Business Unit Head", "Category Owner", "PMO Lead", "IT/Data", "External Advisor"]
_RACI_TASKS = [
    "Approve initiative mandate",
    "Provide spend data access",
    "Validate assumptions & P10/P50/P90",
    "Execute savings actions",
    "Track actuals vs. target",
    "Escalate at-risk initiatives",
    "Sign-off Gate-2 (AQS ≥ 0.65)",
    "Monthly MOR pack review",
]
_RACI_MATRIX = [
    ["A", "I", "C", "R", "I", "C"],
    ["I", "R", "C", "I", "C", "I"],
    ["C", "C", "R", "I", "I", "A"],
    ["I", "I", "C", "R", "I", "I"],
    ["I", "I", "C", "C", "R", "I"],
    ["A", "C", "R", "R", "I", "I"],
    ["A", "C", "R", "I", "I", "C"],
    ["A", "R", "I", "C", "I", "I"],
]


_BENCHMARK_DISCLAIMER = (
    "Benchmark data sourced from internal calibration models. "
    "Figures are illustrative and based on sector-level heuristics — not licensed third-party data "
    "(e.g. Gartner, Hackett Group, NASSCOM). "
    "All savings estimates should be independently validated before client presentation or board submission."
)

_DEFAULT_DISCLAIMER = (
    "Savings figures are model outputs based on spend profiling and sector benchmarks. "
    "P10/P50/P90 ranges represent scenario bounds, not guaranteed outcomes. "
    "All numbers should be reviewed and validated by a qualified FP&A professional prior to use."
)


def build_pmo_data(
    pipeline_summary: Dict[str, Any],
    initiatives: List[Dict[str, Any]],
    *,
    start_date: Optional[date] = None,
    company_name: str = "Client",
    engagement_weeks: int = 12,
    benchmark_disclaimer: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Build structured PMO data from pipeline and initiative lists.
    Returns JSON-serialisable dict.
    """
    anchor = start_date or date.today()

    tracker_rows = []
    for i, init in enumerate(initiatives):
        cid = init.get("category_id") or init.get("initiative_id") or f"INI-{i + 1:03d}"
        p50_cr = round(float(init.get("p50") or 0) / 1e7, 1)
        committed = float(init.get("committed_savings") or 0)
        status = init.get("status") or ("on_track" if p50_cr > 0 else "identified")
        wave = 1 if i < 3 else (2 if i < 6 else 3)
        start_offset = (wave - 1) * 4
        target_date = anchor + timedelta(weeks=start_offset + 4)
        tracker_rows.append({
            "initiative_id": cid,
            "initiative_name": init.get("category_name") or cid,
            "wave": wave,
            "owner": init.get("owner") or "TBD",
            "p50_savings_cr": p50_cr,
            "committed_cr": round(committed / 1e7, 1),
            "status": status,
            "target_date": target_date.isoformat(),
            "ftc_cr": round(float(init.get("forecast_to_complete") or p50_cr), 1),
            "variance_cr": round(p50_cr - float(init.get("forecast_to_complete") or p50_cr), 1),
        })

    # Milestone calendar — 12-week engagement milestones
    milestones = []
    gate_weeks = [3, 6, 9, 12]
    gate_labels = ["Gate-1: Diagnostic sign-off", "Gate-2: Initiative mandate", "Gate-3: Wave-1 delivery", "Gate-4: Programme close"]
    for wk, label in zip(gate_weeks, gate_labels):
        milestones.append({
            "week": wk,
            "date": (anchor + timedelta(weeks=wk)).isoformat(),
            "milestone": label,
            "deliverable": ["CFO brief", "Board deck", "Initiative mandates", "MOR pack + tear-down"][gate_weeks.index(wk)],
        })

    # FTC variance report
    total_p50 = sum(cast(float, r["p50_savings_cr"]) for r in tracker_rows)
    total_ftc = sum(cast(float, r["ftc_cr"]) for r in tracker_rows)
    total_variance = round(total_p50 - total_ftc, 1)
    at_risk = [r for r in tracker_rows if cast(float, r["variance_cr"]) < -0.5]

    disclaimer_text = benchmark_disclaimer or _DEFAULT_DISCLAIMER

    return {
        "type": "pmo_toolkit",
        "generated_on": str(date.today()),
        "company": company_name,
        "initiative_tracker": tracker_rows,
        "raci_matrix": {
            "roles": _RACI_ROLES,
            "tasks": _RACI_TASKS,
            "matrix": _RACI_MATRIX,
        },
        "milestone_calendar": milestones,
        "ftc_report": {
            "total_p50_cr": round(total_p50, 1),
            "total_ftc_cr": round(total_ftc, 1),
            "total_variance_cr": total_variance,
            "at_risk_initiatives": [r["initiative_name"] for r in at_risk],
            "at_risk_count": len(at_risk),
        },
        "disclaimer": disclaimer_text,
    }


def export_pmo_xlsx(pmo: Dict[str, Any], filename: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename

    if not _XLSX_AVAILABLE:
        path.write_text(f"[XLSX stub — openpyxl not installed]\n{pmo['generated_on']}")
        return path

    wb = openpyxl.Workbook()

    # --- Sheet 1: Initiative Tracker ---
    ws1 = wb.active
    ws1.title = "Initiative Tracker"
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="003366")
    hdr_cols = ["ID", "Initiative", "Wave", "Owner", "P50 (₹Cr)", "Committed (₹Cr)", "Status", "Target Date", "FTC (₹Cr)", "Variance (₹Cr)"]
    for j, h in enumerate(hdr_cols, 1):
        cell = ws1.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(pmo["initiative_tracker"], 2):
        ws1.cell(row=i, column=1, value=row["initiative_id"])
        ws1.cell(row=i, column=2, value=row["initiative_name"])
        ws1.cell(row=i, column=3, value=row["wave"])
        ws1.cell(row=i, column=4, value=row["owner"])
        ws1.cell(row=i, column=5, value=row["p50_savings_cr"])
        ws1.cell(row=i, column=6, value=row["committed_cr"])
        ws1.cell(row=i, column=7, value=row["status"])
        ws1.cell(row=i, column=8, value=row["target_date"])
        ws1.cell(row=i, column=9, value=row["ftc_cr"])
        ws1.cell(row=i, column=10, value=row["variance_cr"])

    # --- Sheet 2: RACI ---
    ws2 = wb.create_sheet("RACI Matrix")
    raci = pmo["raci_matrix"]
    ws2.cell(row=1, column=1, value="Task \\ Role").font = hdr_font
    for j, role in enumerate(raci["roles"], 2):
        ws2.cell(row=1, column=j, value=role).font = hdr_font
    for i, (task, row) in enumerate(zip(raci["tasks"], raci["matrix"]), 2):
        ws2.cell(row=i, column=1, value=task)
        for j, val in enumerate(row, 2):
            ws2.cell(row=i, column=j, value=val)

    # --- Sheet 3: Milestone Calendar ---
    ws3 = wb.create_sheet("Milestones")
    for j, h in enumerate(["Week", "Date", "Milestone", "Deliverable"], 1):
        ws3.cell(row=1, column=j, value=h).font = hdr_font
    for i, ms in enumerate(pmo["milestone_calendar"], 2):
        ws3.cell(row=i, column=1, value=ms["week"])
        ws3.cell(row=i, column=2, value=ms["date"])
        ws3.cell(row=i, column=3, value=ms["milestone"])
        ws3.cell(row=i, column=4, value=ms["deliverable"])

    # --- Sheet 4: FTC Report ---
    ws4 = wb.create_sheet("FTC Report")
    ftc = pmo["ftc_report"]
    ws4.cell(row=1, column=1, value="Metric").font = hdr_font
    ws4.cell(row=1, column=2, value="Value").font = hdr_font
    rows = [
        ("Total P50 Savings (₹ Cr)", ftc["total_p50_cr"]),
        ("Total FTC (₹ Cr)", ftc["total_ftc_cr"]),
        ("Total Variance (₹ Cr)", ftc["total_variance_cr"]),
        ("At-Risk Initiative Count", ftc["at_risk_count"]),
        ("At-Risk Initiatives", ", ".join(ftc["at_risk_initiatives"])),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws4.cell(row=i, column=1, value=k)
        ws4.cell(row=i, column=2, value=v)

    # --- Sheet 5: Disclaimer ---
    ws5 = wb.create_sheet("Disclaimer")
    ws5.cell(row=1, column=1, value="Disclaimer").font = Font(bold=True, size=13)
    ws5.cell(row=2, column=1, value=pmo.get("disclaimer", _DEFAULT_DISCLAIMER))
    ws5.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws5.column_dimensions["A"].width = 100
    ws5.cell(row=4, column=1, value=f"Generated: {pmo.get('generated_on', str(date.today()))}")
    ws5.cell(row=5, column=1, value=f"Company: {pmo.get('company', 'Client')}")

    wb.save(path)
    return path
