from __future__ import annotations

import hashlib
import io
import json
import re
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import logger
from app.models import NormalizedSpendLine

# Row-count guardrails — enforced in parse_spend_file before any processing.
_ROW_WARN = 100_000
_ROW_CHUNK = 200_000
_ROW_MAX = 500_000

# ---------------------------------------------------------------------------
# Optional: regional language OCR (pytesseract + Tesseract system binary)
# Falls back gracefully if pytesseract is not installed or Tesseract is absent.
# Supported lang codes: hin (Hindi), tam (Tamil), tel (Telugu), kan (Kannada),
# mar (Marathi), ben (Bengali), guj (Gujarati), pan (Punjabi).
# ---------------------------------------------------------------------------
try:
    import pytesseract as _pytesseract
    from PIL import Image as _PILImage
    _OCR_AVAILABLE = True
except ImportError:
    _OCR_AVAILABLE = False

# Optional: indic-transliteration for script-agnostic text normalisation.
# Used by vendor name dedup to convert Devanagari/Tamil/Telugu → Latin.
try:
    from indic_transliteration import sanscript as _sanscript
    from indic_transliteration.sanscript import transliterate as _transliterate
    _TRANSLITERATION_AVAILABLE = True
except ImportError:
    _TRANSLITERATION_AVAILABLE = False

# Tesseract language pack → ISO script mapping
_LANG_SCRIPTS = {
    "hin": "DEVANAGARI",
    "mar": "DEVANAGARI",
    "tam": "TAMIL",
    "tel": "TELUGU",
    "kan": "KANNADA",
    "ben": "BENGALI",
    "guj": "GUJARATI",
    "pan": "GURMUKHI",
}


_AMOUNT_HEADER_TOKENS = ("amount", "spend", "total", "cost", "value", "inr", "usd")
_SUPPLIER_HEADER_TOKENS = ("supplier", "vendor", "payee", "party", "name")
_DATE_HEADER_TOKENS = ("date", "invoice", "month", "period", "fy")
_CATEGORY_HEADER_TOKENS = ("category", "gl", "account", "cost type", "expense")
_LEDGER_NAME_TOKENS = ("raw", "data", "spend", "detail", "ledger", "transaction", "ap ", "extract", "vendor", "line item")
_SKIP_NAME_TOKENS = ("dashboard", "summary", "cover", "chart", "readme", "index", "toc")


@dataclass
class SheetScore:
    sheet_name: str
    score: float
    header_row: int
    inferred_role: str
    row_count: int
    has_amount_col: bool
    has_supplier_col: bool
    reason: str = ""


def _header_match_score(cell_values: List[str]) -> Tuple[float, bool, bool]:
    """Score a candidate header row; return (score, has_amount, has_supplier)."""
    joined = " ".join(v.lower() for v in cell_values if v)
    has_amount = any(t in joined for t in _AMOUNT_HEADER_TOKENS)
    has_supplier = any(t in joined for t in _SUPPLIER_HEADER_TOKENS)
    has_date = any(t in joined for t in _DATE_HEADER_TOKENS)
    has_category = any(t in joined for t in _CATEGORY_HEADER_TOKENS)
    score = 0.0
    if has_amount:
        score += 3.0
    if has_supplier:
        score += 2.0
    if has_date:
        score += 0.5
    if has_category:
        score += 0.5
    return score, has_amount, has_supplier


def _sniff_header_row_openpyxl(ws: Any, max_scan: int = 15) -> Tuple[int, float, bool, bool]:
    best_row = 0
    best_score = 0.0
    best_amount = False
    best_supplier = False
    for r in range(1, min(max_scan, ws.max_row or 1) + 1):
        vals = [
            str(ws.cell(row=r, column=c).value or "").strip()
            for c in range(1, min(ws.max_column or 1, 40) + 1)
        ]
        if not any(vals):
            continue
        score, has_amount, has_supplier = _header_match_score(vals)
        if score > best_score:
            best_score = score
            best_row = r - 1  # pandas header index (0-based)
            best_amount = has_amount
            best_supplier = has_supplier
    return best_row, best_score, best_amount, best_supplier


def _sheet_name_bonus(name: str) -> Tuple[float, str]:
    low = name.lower()
    bonus = 0.0
    role = "unknown"
    if any(t in low for t in _SKIP_NAME_TOKENS):
        bonus -= 4.0
        role = "summary"
    if any(t in low for t in _LEDGER_NAME_TOKENS):
        bonus += 2.5
        role = "transaction_ledger"
    if any(t in low for t in ("assumption", "input", "driver")):
        role = "assumptions"
    if any(t in low for t in ("scenario", "forecast", "budget", "opex build")):
        role = "timeseries"
    return bonus, role


def score_workbook_sheets(file_path: Path) -> List[SheetScore]:
    """Rank worksheets for transactional spend ingestion (highest score first)."""
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        try:
            frame = pd.read_csv(file_path, nrows=5)
            cols = [str(c) for c in frame.columns]
            hscore, has_amount, has_supplier = _header_match_score(cols)
            return [
                SheetScore(
                    sheet_name=file_path.stem,
                    score=hscore + (2.0 if has_amount and has_supplier else 0),
                    header_row=0,
                    inferred_role="transaction_ledger",
                    row_count=0,
                    has_amount_col=has_amount,
                    has_supplier_col=has_supplier,
                    reason="single CSV file",
                )
            ]
        except Exception:
            return []

    if suffix not in (".xlsx", ".xls"):
        return []

    scores: List[SheetScore] = []
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return []

    for ws in wb.worksheets:
        name = ws.title
        header_row, hscore, has_amount, has_supplier = _sniff_header_row_openpyxl(ws)
        name_bonus, inferred_role = _sheet_name_bonus(name)
        row_count = int(ws.max_row or 0)
        data_rows = max(0, row_count - header_row - 1)

        score = hscore + name_bonus
        if has_amount and has_supplier:
            score += 2.0
        if data_rows >= 10:
            score += min(3.0, data_rows / 500.0)
        elif data_rows < 5:
            score -= 2.0

        period_hits = 0
        if ws.max_row and ws.max_row >= 2:
            row1 = [
                str(ws.cell(row=1, column=c).value or "")
                for c in range(1, min(ws.max_column or 1, 20) + 1)
            ]
            period_hits = sum(1 for v in row1 if _looks_like_period_header(v))
        if period_hits >= 3 and hscore < 4:
            inferred_role = "timeseries"
            score += 1.0

        reason_parts = []
        if has_amount:
            reason_parts.append("amount column")
        if has_supplier:
            reason_parts.append("supplier column")
        if name_bonus > 0:
            reason_parts.append("sheet name signal")

        scores.append(
            SheetScore(
                sheet_name=name,
                score=round(score, 2),
                header_row=header_row,
                inferred_role=inferred_role,
                row_count=row_count,
                has_amount_col=has_amount,
                has_supplier_col=has_supplier,
                reason=", ".join(reason_parts) or "low signal",
            )
        )

    scores.sort(key=lambda s: s.score, reverse=True)
    return scores


def workbook_schema_hints(file_path: Path) -> Dict[str, Any]:
    """Summary for upload manifest schema.workbook."""
    ranked = score_workbook_sheets(file_path)
    selected = ranked[0] if ranked else None
    return {
        "sheet_count": len(ranked),
        "sheet_names": [s.sheet_name for s in ranked],
        "selected_sheet": selected.sheet_name if selected else None,
        "selected_header_row": selected.header_row if selected else 0,
        "sheet_scores": [
            {
                "sheet": s.sheet_name,
                "score": s.score,
                "role": s.inferred_role,
                "header_row": s.header_row,
                "row_count": s.row_count,
            }
            for s in ranked[:5]
        ],
        "planning_signal_confidence": min(1.0, max((s.score for s in ranked), default=0.0) / 8.0),
    }


def _read_tabular(
    file_path: Path,
    sheet_name: str | None = None,
    header_row: int = 0,
) -> pd.DataFrame:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(file_path)
    if sheet_name is not None:
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    ranked = score_workbook_sheets(file_path)
    if ranked and ranked[0].score > 0 and ranked[0].has_amount_col:
        best = ranked[0]
        return pd.read_excel(file_path, sheet_name=best.sheet_name, header=best.header_row)
    return pd.read_excel(file_path)


def _best_column_match(columns: List[str], candidates: List[str], default: str = "") -> str:
    lowered = {c.lower(): c for c in columns}
    for candidate in candidates:
        if candidate in lowered:
            return lowered[candidate]
    for c in columns:
        for candidate in candidates:
            if candidate in c.lower():
                return c
    return default


# Headers that look monetary but are dimensions (Spend_Category, Invoice_ID, …).
_AMOUNT_NEGATIVE_TOKENS = (
    "category",
    "type",
    "id",
    "name",
    "status",
    "department",
    "region",
    "vendor",
    "supplier",
    "date",
    "invoice",
    "memo",
    "description",
    "dept",
    "business unit",
    "geo",
    "location",
    "code",
    "period",
    "month",
    "quarter",
    "scenario",
)


def _numeric_valid_ratio(series: pd.Series) -> float:
    """Share of cells that coerce to a finite number."""
    if series is None or len(series) == 0:
        return 0.0
    numeric = pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False), errors="coerce"
    )
    return float(numeric.notna().sum()) / max(len(series), 1)


def _column_has_negative_amount_token(col: str, has_amount_token: bool) -> bool:
    """True when column name looks like metadata, not a monetary field."""
    if has_amount_token:
        return False
    low = col.lower().strip().replace("_", " ")
    if low.startswith("unnamed"):
        return False
    for token in _AMOUNT_NEGATIVE_TOKENS:
        if re.search(rf"\b{re.escape(token)}\b", low):
            return True
    return False


def _best_numeric_amount_column(
    columns: List[str],
    frame: pd.DataFrame,
    exclude: Optional[set[str]] = None,
    min_ratio: float = 0.25,
) -> str:
    """Pick the column with the strongest numeric density (P&L period columns)."""
    skip = exclude or set()
    best_col = ""
    best_ratio = 0.0
    for col in columns:
        if col in skip or col not in frame.columns:
            continue
        ratio = _numeric_valid_ratio(frame[col])
        if ratio > best_ratio:
            best_ratio = ratio
            best_col = col
    return best_col if best_ratio >= min_ratio else ""


def _best_pl_description_column(
    columns: List[str],
    frame: pd.DataFrame,
    exclude: Optional[set[str]] = None,
) -> str:
    """Pick a line-item label column for hierarchical expense / P&L tables."""
    skip = exclude or set()
    best_col = ""
    best_score = 0.0
    for col in columns:
        if col in skip or col not in frame.columns:
            continue
        series = frame[col]
        if _numeric_valid_ratio(series) > 0.3:
            continue
        non_empty = series.dropna().astype(str).str.strip()
        non_empty = non_empty[(non_empty != "") & (non_empty.str.lower() != "nan")]
        if len(non_empty) < 2:
            continue
        fill_ratio = len(non_empty) / max(len(series), 1)
        avg_len = float(non_empty.str.len().mean())
        score = fill_ratio * 5.0 + min(avg_len / 20.0, 3.0)
        if score > best_score:
            best_score = score
            best_col = col
    return best_col


def _apply_hierarchical_expense_column_mapping(
    columns: List[str],
    frame: pd.DataFrame,
    amount_col: str,
    desc_col: str,
) -> Tuple[str, str, Optional[str]]:
    """Repair amount/description mapping for P&L-style sheets (labels + period amounts)."""
    amount_ratio = _numeric_valid_ratio(frame[amount_col]) if amount_col and amount_col in frame.columns else 0.0
    numeric_col = _best_numeric_amount_column(columns, frame)
    numeric_ratio = _numeric_valid_ratio(frame[numeric_col]) if numeric_col else 0.0

    if not numeric_col or numeric_ratio < 0.25:
        return amount_col, desc_col, None
    if numeric_ratio <= amount_ratio + 0.15:
        return amount_col, desc_col, None

    label_col = _best_pl_description_column(columns, frame, exclude={numeric_col})
    note = (
        f"Hierarchical expense layout: amounts from '{numeric_col}'"
        + (f", line items from '{label_col}'." if label_col else ".")
    )
    return numeric_col, label_col or desc_col, note


def _column_amount_score(col: str, series: Optional[pd.Series] = None) -> float:
    """Score how likely a column holds numeric spend amounts (higher = better)."""
    low = col.lower().strip().replace("_", " ")
    if "spend" in low and any(t in low for t in ("category", "cat", "class", "type")):
        return -1.0
    has_amount_token = any(t in low for t in ("amount", "value", "total", "cost", "price", "fee"))
    if _column_has_negative_amount_token(col, has_amount_token):
        return -1.0

    score = 0.0
    if low in ("amount", "value", "cost", "total"):
        score += 10.0
    elif low.startswith("amount") or " amount" in low or low.endswith(" amount"):
        score += 9.0
        if any(ccy in low for ccy in ("usd", "inr", "eur", "gbp", "sgd", "aed")):
            score += 3.0
    elif "amount" in low:
        score += 8.0
    elif "total" in low and "category" not in low:
        score += 6.0
    elif "cost" in low and "center" not in low and "centre" not in low:
        score += 5.0
    elif "value" in low:
        score += 5.0
    elif "spend" in low:
        score += 2.0

    if series is not None and len(series) > 0:
        valid_ratio = _numeric_valid_ratio(series)
        if has_amount_token and valid_ratio < 0.2:
            # e.g. "Cost of Goods Sold" column holding line-item labels, not amounts
            return valid_ratio * 12.0
        if valid_ratio > 0.8:
            score += 8.0
        elif valid_ratio > 0.5:
            score += 4.0
        elif valid_ratio > 0.25:
            score += 2.0
        numeric = pd.to_numeric(
            series.astype(str).str.replace(",", "", regex=False), errors="coerce"
        )
        if numeric.notna().any() and float(numeric.fillna(0).median()) > 0:
            score += 2.0
    return score


def _best_amount_column_match(
    columns: List[str],
    frame: Optional[pd.DataFrame] = None,
    default: str = "",
) -> str:
    """Pick the monetary amount column; avoids Spend_Category-style false positives."""
    best_col = default
    best_score = 0.0
    for col in columns:
        series = frame[col] if frame is not None and col in frame.columns else None
        score = _column_amount_score(col, series)
        if score > best_score:
            best_score = score
            best_col = col
    return best_col


def _infer_currency_from_column_name(col_name: str) -> Optional[str]:
    low = col_name.lower()
    for code in ("usd", "inr", "eur", "gbp", "sgd", "aed", "chf", "jpy"):
        if code in low:
            return code.upper()
    if "rupee" in low or "₹" in col_name:
        return "INR"
    if "$" in col_name:
        return "USD"
    return None


def _classify(description: str, supplier: str, taxonomy: Dict) -> Tuple[str, str]:
    content = f"{description} {supplier}".lower()
    best = ("OTHER", "Other / Unclassified", 0)
    for cat in taxonomy.get("categories", []):
        hits = sum(1 for kw in cat.get("keywords", []) if kw.lower() in content)
        if hits > best[2]:
            best = (cat.get("id", "OTHER"), cat.get("name", "Other / Unclassified"), hits)
    return best[0], best[1]


def _classify_by_gl(gl_code: str, taxonomy: Dict) -> Tuple[str, str] | None:
    """Map a GL code to a taxonomy category using gl_code_ranges in taxonomy."""
    if not gl_code:
        return None
    gl_ranges = taxonomy.get("gl_code_ranges", {})
    stripped = re.sub(r"\D", "", gl_code)
    if not stripped:
        logger.warning("gl_code_no_digits gl_code=%r; skipping GL classification", gl_code)
        return None
    try:
        code_int = int(stripped)
    except (ValueError, TypeError):
        return None
    for range_key, mapping in gl_ranges.items():
        try:
            lo, hi = range_key.split("-")
            if int(lo) <= code_int <= int(hi):
                return mapping.get("category_id", "OTHER"), mapping.get("category_name", "Other / Unclassified")
        except Exception:
            continue
    return None


def _derive_fiscal_period(date_str: str) -> Tuple[int | None, str | None]:
    """Return (fiscal_year, fiscal_period) from a date string.

    fiscal_period format: "YYYY-Qn" for quarterly or "YYYY-MM" for monthly.
    We prefer monthly precision when available.
    """
    if not date_str or date_str in ("", "nan", "NaT", "None"):
        return None, None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m", "%m/%Y", "%b %Y", "%B %Y"):
        try:
            dt = datetime.strptime(str(date_str).strip()[:10], fmt)
            fy = dt.year
            fp = f"{dt.year}-{dt.month:02d}"
            return fy, fp
        except ValueError:
            continue
    # Try pandas as fallback for exotic formats
    try:
        dt = pd.to_datetime(date_str)
        fy = dt.year
        fp = f"{dt.year}-{dt.month:02d}"
        return fy, fp
    except Exception:
        return None, None


def _detect_amount_type(value: str) -> str:
    """Infer amount_type from a column name or cell value."""
    low = str(value).lower()
    if any(k in low for k in ("budget", "plan", "planned")):
        return "budget"
    if any(k in low for k in ("forecast", "fcast", "fcst", "projection")):
        return "forecast"
    if "accrual" in low:
        return "accrual"
    return "actual"


def _parse_fx_rate(value: Any) -> float:
    try:
        v = float(value)
        if v > 0:
            return v
        logger.warning("fx_rate_invalid value=%r; defaulting to 1.0", value)
        return 1.0
    except (TypeError, ValueError):
        logger.warning("fx_rate_parse_failed value=%r; defaulting to 1.0", value)
        return 1.0


def _parse_gst_treatment(value: Any) -> str | None:
    """Normalise a raw GST treatment cell to canonical tag."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if any(k in s for k in ("ineligible", "blocked", "exempt", "nil rated", "non-itc")):
        return "ineligible"
    if any(k in s for k in ("rcm", "reverse charge")):
        return "rcm"
    if any(k in s for k in ("inverted", "refund")):
        return "inverted_duty"
    if any(k in s for k in ("eligible", "itc", "input tax")):
        return "itc_eligible"
    return None


def _parse_related_party(value: Any) -> bool:
    """Return True when the cell indicates a related-party / intercompany transaction."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    s = str(value).strip().lower()
    return s in {"yes", "y", "true", "1", "intercompany", "related", "intragroup", "intra-group"}


def _parse_lease_treatment(value: Any) -> str | None:
    """Normalise a raw lease treatment cell."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if any(k in s for k in ("operating", "op lease", "ind as 116")):
        return "operating_ind_as_116"
    if any(k in s for k in ("finance", "financial", "capital")):
        return "finance"
    if any(k in s for k in ("short", "low value")):
        return "short_term"
    return None


def _parse_payment_terms(value: Any) -> int | None:
    """Extract numeric days from values like 'Net 30', '45', 'Net-60'."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).lower().strip()
    m = re.search(r"\d+", s)
    if m:
        days = int(m.group())
        return days if 0 < days <= 365 else None
    return None


_GSTIN_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b")
_PAN_RE = re.compile(r"\b[A-Z]{5}[0-9]{4}[A-Z]\b")

_CAPEX_KW = ("capital", "capex", "equipment", "machinery", "plant", "property", "construction",
              "building", "vehicle", "computer hardware", "server", "fixed asset", "asset purchase")
_LEASE_KW = ("rent", "lease", "operating lease", "ind as 116", "licence fee", "licensing fee",
              "leasehold")
_STATUTORY_KW = ("gst", "tds", "customs duty", "excise", "cess", " pf ", "esic", "epf",
                 "provident fund", "professional tax", "stamp duty", "court fee",
                 "penalty", "fine", "statutory levy", "import duty")


def _extract_gstin(text: str) -> Optional[str]:
    """Extract the first valid-format GSTIN from a text string."""
    if not text:
        return None
    m = _GSTIN_RE.search(str(text).upper().strip())
    return m.group(0) if m else None


def _extract_pan(text: str) -> Optional[str]:
    """Extract the first valid-format PAN from a text string."""
    if not text:
        return None
    m = _PAN_RE.search(str(text).upper().strip())
    return m.group(0) if m else None


def _classify_spend_type(
    description: str,
    supplier: str,
    gl_code: Optional[str],
    related_party: bool,
    lease_treatment: Optional[str],
) -> str:
    """Classify a spend line into opex / capex / lease / statutory / intercompany."""
    if related_party:
        return "intercompany"
    if lease_treatment in ("operating_ind_as_116", "finance", "short_term"):
        return "lease"
    if gl_code:
        _stripped = re.sub(r"\D", "", gl_code)
        if _stripped:
            try:
                code_int = int(_stripped)
                # Capital/fixed-asset GL range (typical Indian SAP COA)
                if 1000 <= code_int <= 1999 or 10000 <= code_int <= 19999:
                    return "capex"
                # Statutory/tax payable GL range
                if 2300 <= code_int <= 2699 or 23000 <= code_int <= 26999:
                    return "statutory"
            except (ValueError, TypeError):
                pass
    content = f"{description} {supplier}".lower()
    if any(kw in content for kw in _LEASE_KW):
        return "lease"
    if any(kw in content for kw in _CAPEX_KW):
        return "capex"
    if any(kw in content for kw in _STATUTORY_KW):
        return "statutory"
    return "opex"


def _compute_addressable(spend_type: str, related_party: bool) -> bool:
    """Statutory spend and intercompany eliminations are never addressable."""
    if related_party or spend_type == "intercompany":
        return False
    if spend_type == "statutory":
        return False
    return True  # opex / capex / lease all carry addressability levers


def _parse_vendor_category(value: Any) -> Optional[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().lower()
    if any(k in s for k in ("msme", "sme", "small", "micro", "medium enterprise")):
        return "msme"
    if any(k in s for k in ("startup", "start-up", "start up")):
        return "startup"
    if any(k in s for k in ("foreign", "overseas", "international", "mnc", "global")):
        return "foreign"
    if any(k in s for k in ("large", "enterprise", "corporate")):
        return "large"
    return None


def _parse_msme_flag(value: Any) -> Optional[bool]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().lower()
    if s in {"yes", "y", "true", "1", "msme", "sme"}:
        return True
    if s in {"no", "n", "false", "0", "large", "non-msme"}:
        return False
    return None


def _compute_file_hash(file_path: Path) -> str:
    """SHA-256 of the file bytes — used for source dedup."""
    h = hashlib.sha256()
    with open(file_path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65_536), b""):
            h.update(chunk)
    return h.hexdigest()


def infer_tabular_schema(file_path: Path) -> Dict[str, Any]:
    frame = _read_tabular(file_path)
    columns = [str(c) for c in frame.columns]
    role_map = {
        "amount": _best_amount_column_match(columns, frame),
        "supplier": _best_column_match(columns, ["supplier", "vendor", "payee"]),
        "description": _best_column_match(columns, ["description", "memo", "line item", "item"]),
        "business_unit": _best_column_match(columns, ["bu", "business unit", "department"]),
        "geo": _best_column_match(columns, ["country", "geo", "region", "location"]),
        "date": _best_column_match(columns, ["date", "invoice date", "month"]),
        # FP&A additions
        "gl_code": _best_column_match(columns, ["gl code", "gl_code", "account", "account code", "gl account"]),
        "cost_center": _best_column_match(columns, ["cost center", "cost_center", "cc", "cost centre"]),
        "currency": _best_column_match(columns, ["currency", "ccy", "curr"]),
        "fx_rate": _best_column_match(columns, ["fx rate", "fx_rate", "exchange rate", "rate"]),
        "amount_type": _best_column_match(columns, ["type", "amount type", "data type", "scenario"]),
        "payment_terms": _best_column_match(columns, ["payment terms", "payment_terms", "terms", "net days", "dpo"]),
        # India additions (v2.0)
        "gst_treatment": _best_column_match(columns, ["gst treatment", "gst_treatment", "itc", "gst type", "tax treatment"]),
        "gstin": _best_column_match(columns, ["gstin", "gst number", "gst no", "vendor gstin", "supplier gstin"]),
        "lease_treatment": _best_column_match(columns, ["lease treatment", "lease_treatment", "lease type", "ind as 116"]),
        "related_party": _best_column_match(columns, ["related party", "related_party", "intercompany", "intra group", "intragroup"]),
        "legal_entity": _best_column_match(columns, ["legal entity", "legal_entity", "entity", "company code", "co code"]),
        # Source lineage (v2.2)
        "source_system": _best_column_match(columns, ["source system", "source_system", "source", "erp", "system name"]),
        "source_record_id": _best_column_match(columns, ["source record", "source record id", "record id", "source id", "pk", "primary key"]),
        # Vendor master (v2.2)
        "vendor_pan": _best_column_match(columns, ["vendor pan", "pan", "pan number", "pan no", "supplier pan"]),
        "vendor_msme": _best_column_match(columns, ["msme", "msme flag", "vendor msme", "sme", "is msme"]),
        "vendor_category": _best_column_match(columns, ["vendor category", "vendor_category", "supplier category", "vendor type", "supplier type"]),
        # Spend classification (v2.2)
        "spend_type": _best_column_match(columns, ["spend type", "spend_type", "cost type", "opex capex", "asset type", "expenditure type"]),
    }
    inferred = []
    for column in columns:
        series = frame[column]
        sample_values = [str(x) for x in series.dropna().head(3).tolist()]
        dtype = str(series.dtype)
        semantic_role = ""
        for role, matched_col in role_map.items():
            if matched_col == column:
                semantic_role = role
                break
        inferred.append(
            {
                "name": column,
                "dtype": dtype,
                "semantic_role": semantic_role or "other",
                "sample_values": sample_values,
                "null_ratio": float(series.isna().mean()) if len(series) else 0.0,
            }
        )

    result = {
        "file_name": file_path.name,
        "rows": int(len(frame)),
        "columns": inferred,
        "semantic_map": role_map,
    }
    if file_path.suffix.lower() in (".xlsx", ".xls"):
        hints = workbook_schema_hints(file_path)
        result["workbook"] = {
            **_infer_workbook_summary(file_path),
            **hints,
        }
    return result


def _infer_workbook_summary(file_path: Path) -> Dict[str, Any]:
    try:
        wb = load_workbook(file_path, data_only=False, read_only=True)
    except Exception:
        return {"sheet_count": 0, "sheet_names": [], "planning_signal_confidence": 0.0, "role_hints": []}

    sheet_names = [ws.title for ws in wb.worksheets]
    role_hints: List[str] = []
    confidence = 0.0
    planning_name_tokens = (
        "assumption",
        "input",
        "driver",
        "summary",
        "scenario",
        "plan",
        "forecast",
        "budget",
        "p&l",
    )
    for name in sheet_names:
        low = name.lower()
        if any(tok in low for tok in planning_name_tokens):
            role_hints.append(name)
    if len(sheet_names) > 1:
        confidence += 0.35
    if role_hints:
        confidence += min(0.4, 0.08 * len(role_hints))
    # quick period-header scan on first two rows
    period_hits = 0
    for ws in wb.worksheets[:5]:
        for r in (1, 2):
            row_values = [str(ws.cell(row=r, column=c).value or "") for c in range(1, min(ws.max_column, 15) + 1)]
            hits = sum(1 for val in row_values if _looks_like_period_header(val))
            if hits >= 3:
                period_hits += 1
                break
    if period_hits:
        confidence += min(0.25, 0.08 * period_hits)

    return {
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "planning_signal_confidence": max(0.0, min(1.0, confidence)),
        "role_hints": role_hints[:10],
    }


def _looks_like_period_header(value: str) -> bool:
    s = (value or "").strip().lower()
    if not s:
        return False
    patterns = (
        r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\s\-_/]?\d{2,4}$",
        r"^q[1-4][\s\-_/]?\d{2,4}$",
        r"^fy\d{2,4}$",
        r"^\d{4}[\-_/]\d{1,2}$",
        r"^\d{4}$",
    )
    return any(re.match(p, s) for p in patterns)


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        if isinstance(value, str):
            cleaned = value.replace(",", "").strip()
            if cleaned in ("", "-", "na", "n/a"):
                return None
            return float(cleaned)
        return float(value)
    except Exception:
        return None


def _parse_planning_workbook(
    file_path: Path,
    taxonomy: Dict,
    workbook_manifest: Dict[str, Any],
    reporting_currency: str,
) -> List[NormalizedSpendLine]:
    out: List[NormalizedSpendLine] = []
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return out

    row_id = 1
    nodes = workbook_manifest.get("sheet_graph", []) if isinstance(workbook_manifest, dict) else []
    for node in nodes:
        role = str(node.get("role") or "unknown")
        if role not in {"timeseries", "scenarios"}:
            continue
        sheet_name = str(node.get("sheet_name") or "")
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        axis = node.get("period_axis") or {}
        if str(axis.get("orientation") or "") != "column":
            continue
        first_period_col = int(axis.get("first_period_col") or 2)
        periods = list(axis.get("periods") or [])
        if not periods:
            # fallback: detect from first row
            periods = [
                str(ws.cell(row=1, column=c).value or "").strip()
                for c in range(first_period_col, min(ws.max_column, first_period_col + 36) + 1)
                if _looks_like_period_header(str(ws.cell(row=1, column=c).value or ""))
            ]
        if not periods:
            continue
        for r in range(2, min(ws.max_row, 2000) + 1):
            label = str(ws.cell(row=r, column=1).value or "").strip()
            if not label:
                continue
            category_id, category_name = _classify(label, "", taxonomy)
            for idx, period in enumerate(periods):
                col = first_period_col + idx
                if col > ws.max_column:
                    break
                amount = _to_float(ws.cell(row=r, column=col).value)
                if amount is None:
                    continue
                amount_type = _detect_amount_type(period)
                fiscal_year, fiscal_period = _derive_fiscal_period(str(period))
                out.append(
                    NormalizedSpendLine(
                        row_id=row_id,
                        supplier="Model",
                        description=label,
                        amount=amount,
                        category_id=category_id,
                        category_name=category_name,
                        business_unit=sheet_name,
                        geo=None,
                        spend_date=None,
                        gl_code=None,
                        cost_center_id=None,
                        currency=reporting_currency,
                        fx_rate_to_reporting=1.0,
                        amount_reporting=amount,
                        amount_type=amount_type,
                        fiscal_year=fiscal_year,
                        fiscal_period=fiscal_period or str(period),
                        payment_terms_days=None,
                    )
                )
                row_id += 1
    return out


def _dataframe_to_spend_lines(
    frame: pd.DataFrame,
    file_path: Path,
    taxonomy: Dict,
    default_amount_type: str = "actual",
    reporting_currency: str = "INR",
    source_system_id: Optional[str] = None,
    row_id_start: int = 1,
) -> Tuple[List[NormalizedSpendLine], Optional[str]]:
    file_hash = _compute_file_hash(file_path)

    # --- Row-count guardrails ---
    n_rows = len(frame)
    if n_rows > _ROW_MAX:
        raise ValueError(
            f"{file_path.name} has {n_rows:,} rows (limit: {_ROW_MAX:,}). "
            "Split by entity or fiscal year and upload each part separately."
        )
    if n_rows > _ROW_CHUNK:
        logger.warning(
            '"large_file_perf_warning rows=%d file=%s — consider splitting for faster processing"',
            n_rows,
            file_path.name,
        )
    elif n_rows > _ROW_WARN:
        logger.warning('"large_file rows=%d file=%s"', n_rows, file_path.name)

    cols = [str(c) for c in frame.columns]
    amount_col = _best_amount_column_match(cols, frame)
    supplier_col = _best_column_match(cols, ["supplier", "vendor", "payee"], "")
    desc_col = _best_column_match(cols, ["description", "memo", "line item", "item"], "")
    amount_col, desc_col, pl_note = _apply_hierarchical_expense_column_mapping(
        cols, frame, amount_col, desc_col
    )
    mapping_note = pl_note
    # Prefer an explicit spend-category column as the line label before the
    # generic "longest text column" heuristic. Otherwise a Region or Department
    # column can shadow Spend_Category as the description on a normal ledger.
    if not desc_col:
        desc_col = _best_column_match(
            cols,
            ["spend category", "spend_category", "expense category", "category"],
            "",
        )
    label_col = _best_pl_description_column(
        cols, frame, exclude={amount_col} if amount_col else set()
    )
    if label_col:
        use_label = not desc_col
        if desc_col and desc_col in frame.columns:
            filled = frame[desc_col].dropna().astype(str).str.strip()
            filled = filled[(filled != "") & (filled.str.lower() != "nan")]
            use_label = len(filled) < max(2, int(len(frame) * 0.15))
        if use_label:
            desc_col = label_col
            if not mapping_note:
                mapping_note = f"Hierarchical expense layout: line items from '{label_col}'."
    amount_currency_hint = _infer_currency_from_column_name(amount_col) if amount_col else None
    bu_col = _best_column_match(cols, ["bu", "business unit", "department"], "")
    geo_col = _best_column_match(cols, ["country", "geo", "region", "location"], "")
    date_col = _best_column_match(cols, ["date", "invoice date", "month"], "")
    # FP&A columns
    gl_col = _best_column_match(cols, ["gl code", "gl_code", "account", "account code", "gl account"], "")
    cc_col = _best_column_match(cols, ["cost center", "cost_center", "cc", "cost centre"], "")
    currency_col = _best_column_match(cols, ["currency", "ccy", "curr"], "")
    fx_col = _best_column_match(cols, ["fx rate", "fx_rate", "exchange rate"], "")
    type_col = _best_column_match(cols, ["type", "amount type", "data type", "scenario"], "")
    terms_col = _best_column_match(cols, ["payment terms", "payment_terms", "terms", "net days", "dpo"], "")
    # India columns (v2.0)
    gst_col = _best_column_match(cols, ["gst treatment", "gst_treatment", "itc", "gst type", "tax treatment"], "")
    gstin_col = _best_column_match(cols, ["gstin", "gst number", "gst no", "vendor gstin", "supplier gstin"], "")
    lease_col = _best_column_match(cols, ["lease treatment", "lease_treatment", "lease type", "ind as 116"], "")
    rp_col = _best_column_match(cols, ["related party", "related_party", "intercompany", "intra group", "intragroup"], "")
    entity_col = _best_column_match(cols, ["legal entity", "legal_entity", "entity", "company code", "co code"], "")
    # Source lineage columns (v2.2)
    src_sys_col = _best_column_match(cols, ["source system", "source_system", "source", "erp", "system name"], "")
    src_rec_col = _best_column_match(cols, ["source record id", "source record", "record id", "source id", "pk"], "")
    # Vendor master columns (v2.2)
    pan_col = _best_column_match(cols, ["vendor pan", "pan", "pan number", "pan no", "supplier pan"], "")
    msme_col = _best_column_match(cols, ["msme", "msme flag", "vendor msme", "sme", "is msme"], "")
    vcat_col = _best_column_match(cols, ["vendor category", "vendor_category", "supplier category", "vendor type"], "")
    # Spend classification column (v2.2)
    stype_col = _best_column_match(cols, ["spend type", "spend_type", "cost type", "opex capex", "expenditure type"], "")

    def _col(name: str, fill: str = "") -> "pd.Series":
        return frame[name].fillna(fill) if name else pd.Series([""] * len(frame))

    amount_s = frame[amount_col] if amount_col else pd.Series([0] * len(frame))
    supplier_s = _col(supplier_col)
    desc_s = _col(desc_col)
    bu_s = _col(bu_col)
    geo_s = _col(geo_col)
    date_s = _col(date_col)
    gl_s = _col(gl_col)
    cc_s = _col(cc_col)
    currency_s = _col(currency_col, reporting_currency)
    fx_s = _col(fx_col, "1.0")
    type_s = _col(type_col, default_amount_type)
    terms_s = _col(terms_col)
    # India series (v2.0)
    gst_s = _col(gst_col)
    gstin_s = _col(gstin_col)
    lease_s = _col(lease_col)
    rp_s = _col(rp_col)
    entity_s = _col(entity_col)
    # Source lineage + vendor master series (v2.2)
    src_sys_s = _col(src_sys_col)
    src_rec_s = _col(src_rec_col)
    pan_s = _col(pan_col)
    msme_s = _col(msme_col)
    vcat_s = _col(vcat_col)
    stype_s = _col(stype_col)

    # --- Vectorized pre-computation (avoids per-row Python overhead on hot paths) ---
    # Amount: pandas coerce-to-numeric is ~10x faster than per-row float() in a loop.
    _amount_numeric = pd.to_numeric(
        amount_s.astype(str).str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0.0)

    out: List[NormalizedSpendLine] = []
    for i in range(len(frame)):
        amount = float(_amount_numeric.iloc[i])
        if amount == 0.0:
            continue

        supplier = str(supplier_s.iloc[i]).strip()
        description = str(desc_s.iloc[i]).strip()
        date_raw = str(date_s.iloc[i]).strip()
        gl_code = str(gl_s.iloc[i]).strip() or None
        cost_center = str(cc_s.iloc[i]).strip() or None
        currency = (
            str(currency_s.iloc[i]).strip()
            or amount_currency_hint
            or reporting_currency
        )
        fx_rate = _parse_fx_rate(fx_s.iloc[i])
        amount_type = _detect_amount_type(str(type_s.iloc[i]).strip()) if type_col else default_amount_type
        payment_terms = _parse_payment_terms(terms_s.iloc[i])

        # GL code is primary classifier; fall back to keyword match
        if gl_code:
            gl_result = _classify_by_gl(gl_code, taxonomy)
        else:
            gl_result = None
        if gl_result:
            category_id, category_name = gl_result
        else:
            category_id, category_name = _classify(description, supplier, taxonomy)

        fiscal_year, fiscal_period = _derive_fiscal_period(date_raw)

        # Compute reporting amount (apply FX only when currency differs)
        if currency.upper() != reporting_currency.upper():
            amount_reporting = amount * fx_rate
        else:
            amount_reporting = amount

        gst_treatment = _parse_gst_treatment(gst_s.iloc[i]) if gst_col else None
        gstin_raw = str(gstin_s.iloc[i]).strip() if gstin_col else ""
        gstin = gstin_raw or None
        # GSTIN extraction: prefer explicit column; fall back to supplier/description text
        vendor_gstin = _extract_gstin(gstin_raw) or _extract_gstin(supplier) or _extract_gstin(description)
        lease_treatment = _parse_lease_treatment(lease_s.iloc[i]) if lease_col else None
        related_party_flag = _parse_related_party(rp_s.iloc[i]) if rp_col else False
        legal_entity_id = str(entity_s.iloc[i]).strip() or None if entity_col else None
        # Source lineage (v2.2)
        row_source_system = str(src_sys_s.iloc[i]).strip() or None if src_sys_col else None
        effective_source_system = row_source_system or source_system_id
        source_record_id = str(src_rec_s.iloc[i]).strip() or None if src_rec_col else None
        # Vendor master enrichment (v2.2)
        vendor_pan = _extract_pan(str(pan_s.iloc[i]).strip()) if pan_col else None
        vendor_msme_flag = _parse_msme_flag(msme_s.iloc[i]) if msme_col else None
        vendor_category = _parse_vendor_category(vcat_s.iloc[i]) if vcat_col else None
        # Spend type: prefer explicit column; compute if absent
        if stype_col and str(stype_s.iloc[i]).strip().lower() in ("opex", "capex", "lease", "statutory", "intercompany"):
            spend_type = str(stype_s.iloc[i]).strip().lower()
        else:
            spend_type = _classify_spend_type(description, supplier, gl_code, related_party_flag, lease_treatment)
        is_addressable = _compute_addressable(spend_type, related_party_flag)
        is_intercompany = True if related_party_flag or spend_type == "intercompany" else None

        out.append(
            NormalizedSpendLine(
                row_id=row_id_start + i,
                supplier=supplier or "Unknown",
                description=description or "N/A",
                amount=amount,
                category_id=category_id,
                category_name=category_name,
                business_unit=str(bu_s.iloc[i]).strip() or None,
                geo=str(geo_s.iloc[i]).strip() or None,
                spend_date=date_raw or None,
                gl_code=gl_code,
                cost_center_id=cost_center,
                currency=currency,
                fx_rate_to_reporting=fx_rate,
                amount_reporting=amount_reporting,
                amount_type=amount_type,
                fiscal_year=fiscal_year,
                fiscal_period=fiscal_period,
                payment_terms_days=payment_terms,
                gst_treatment=gst_treatment,
                gstin=gstin,
                vendor_gstin=vendor_gstin,
                lease_treatment=lease_treatment,
                related_party_flag=related_party_flag,
                legal_entity_id=legal_entity_id,
                # Source lineage
                source_system_id=effective_source_system,
                source_record_id=source_record_id,
                source_file_hash=file_hash,
                is_intercompany=is_intercompany,
                # Vendor master
                vendor_pan=vendor_pan,
                vendor_msme_flag=vendor_msme_flag,
                vendor_category=vendor_category,
                # Classification
                spend_type=spend_type,
                is_addressable=is_addressable,
            )
        )
    return out, mapping_note


def _enrich_ingestion_report(
    report: Dict[str, Any],
    lines: List[NormalizedSpendLine],
    mapping_note: Optional[str] = None,
) -> Dict[str, Any]:
    """Attach spend-quality flags so the UI can warn on zero-total ingests."""
    total_amount = sum(float(line.amount) for line in lines)
    rows_with_amount = sum(1 for line in lines if float(line.amount) != 0.0)
    quality: Dict[str, Any] = {
        "rows_parsed": len(lines),
        "rows_with_amount": rows_with_amount,
        "total_amount": round(total_amount, 2),
        "zero_spend_warning": total_amount == 0.0
        and (len(lines) > 0 or int(report.pop("_empty_parse_rows", 0) or 0) > 0),
    }
    if mapping_note:
        quality["column_mapping_note"] = mapping_note
    report["quality"] = quality
    if quality["zero_spend_warning"]:
        report["warnings"] = [
            "Rows were read from the file but all parsed amounts are zero. "
            "Check that amount columns contain numbers (not only line-item labels)."
        ]
    return report


def _ledger_sheets_from_manifest(workbook_manifest: Dict[str, Any] | None) -> List[str]:
    if not workbook_manifest:
        return []
    names: List[str] = []
    for node in workbook_manifest.get("sheet_graph", []):
        if isinstance(node, dict) and node.get("role") == "transaction_ledger":
            sn = str(node.get("sheet_name") or "").strip()
            if sn:
                names.append(sn)
    return names


def _build_ingestion_skip_list(
    file_path: Path,
    ingested: List[str],
) -> List[Dict[str, str]]:
    skipped: List[Dict[str, str]] = []
    for sc in score_workbook_sheets(file_path):
        if sc.sheet_name in ingested:
            continue
        skipped.append(
            {
                "sheet": sc.sheet_name,
                "role": sc.inferred_role,
                "reason": "non-ledger" if sc.inferred_role in {"summary", "unknown"} else "lower score",
            }
        )
    return skipped


def parse_spend_file_with_report(
    file_path: Path,
    taxonomy: Dict,
    default_amount_type: str = "actual",
    reporting_currency: str = "INR",
    workbook_manifest: Dict[str, Any] | None = None,
    source_system_id: Optional[str] = None,
) -> Tuple[List[NormalizedSpendLine], Dict[str, Any]]:
    """Parse spend file and return normalized lines plus ingestion diagnostics."""
    report: Dict[str, Any] = {
        "source_file": file_path.name,
        "sheets_ingested": [],
        "sheets_skipped": [],
    }
    strategy = str((workbook_manifest or {}).get("ingestion_strategy") or "standard")
    suffix = file_path.suffix.lower()
    ranked = score_workbook_sheets(file_path) if suffix in (".xlsx", ".xls") else []
    score_by_name = {s.sheet_name: s for s in ranked}

    planning_strategies = {
        "timeseries_flatten",
        "scenario_pivot",
        "hybrid",
        "assumptions_extract",
    }
    ledger_from_manifest = _ledger_sheets_from_manifest(workbook_manifest)

    if suffix in (".xlsx", ".xls") and strategy in planning_strategies and strategy != "ledger_standard":
        planning_lines = _parse_planning_workbook(
            file_path=file_path,
            taxonomy=taxonomy,
            workbook_manifest=workbook_manifest or {},
            reporting_currency=reporting_currency,
        )
        if planning_lines:
            for node in (workbook_manifest or {}).get("sheet_graph", []):
                if node.get("role") in {"timeseries", "scenarios"}:
                    report["sheets_ingested"].append(
                        {
                            "sheet": node.get("sheet_name"),
                            "rows": len(planning_lines),
                            "strategy": strategy,
                        }
                    )
            report["sheets_skipped"] = _build_ingestion_skip_list(
                file_path, [x["sheet"] for x in report["sheets_ingested"] if x.get("sheet")]
            )
            return planning_lines, _enrich_ingestion_report(report, planning_lines)

    all_lines: List[NormalizedSpendLine] = []
    ingested_names: List[str] = []
    mapping_notes: List[str] = []

    def _parse_one_sheet(sheet_name: str, strategy_label: str) -> int:
        sc = score_by_name.get(sheet_name)
        header_row = sc.header_row if sc else 0
        frame = _read_tabular(file_path, sheet_name=sheet_name, header_row=header_row)
        n_rows = len(frame)
        if n_rows > _ROW_MAX:
            raise ValueError(
                f"{file_path.name} sheet '{sheet_name}' has {n_rows:,} rows (limit: {_ROW_MAX:,})."
            )
        sid = source_system_id or sheet_name
        lines, mapping_note = _dataframe_to_spend_lines(
            frame,
            file_path,
            taxonomy,
            default_amount_type=default_amount_type,
            reporting_currency=reporting_currency,
            source_system_id=sid,
            row_id_start=len(all_lines) + 1,
        )
        if mapping_note:
            mapping_notes.append(mapping_note)
            report["layout"] = "hierarchical_expense"
        if not lines and n_rows > 0:
            report["_empty_parse_rows"] = int(report.get("_empty_parse_rows", 0)) + n_rows
        if lines:
            all_lines.extend(lines)
            ingested_names.append(sheet_name)
            report["sheets_ingested"].append(
                {"sheet": sheet_name, "rows": len(lines), "strategy": strategy_label}
            )
        return len(lines)

    if suffix in (".xlsx", ".xls"):
        sheets_to_try: List[str] = []
        if ledger_from_manifest:
            sheets_to_try = ledger_from_manifest
        elif strategy == "ledger_standard":
            sheets_to_try = [
                s.sheet_name
                for s in ranked
                if s.inferred_role == "transaction_ledger"
                or (s.has_amount_col and s.has_supplier_col and s.score >= 3.0)
            ]
        if not sheets_to_try and ranked:
            sheets_to_try = [ranked[0].sheet_name]

        for sheet_name in sheets_to_try:
            _parse_one_sheet(sheet_name, "ledger_standard")

        if all_lines:
            report["sheets_skipped"] = _build_ingestion_skip_list(file_path, ingested_names)
            note = mapping_notes[0] if mapping_notes else None
            return all_lines, _enrich_ingestion_report(report, all_lines, note)

    sc = ranked[0] if ranked else None
    if suffix in (".xlsx", ".xls") and sc:
        _parse_one_sheet(sc.sheet_name, "standard")
    else:
        frame = _read_tabular(file_path)
        n_rows = len(frame)
        if n_rows > _ROW_MAX:
            raise ValueError(
                f"{file_path.name} has {n_rows:,} rows (limit: {_ROW_MAX:,}). "
                "Split by entity or fiscal year and upload each part separately."
            )
        lines, mapping_note = _dataframe_to_spend_lines(
            frame,
            file_path,
            taxonomy,
            default_amount_type=default_amount_type,
            reporting_currency=reporting_currency,
            source_system_id=source_system_id,
        )
        if mapping_note:
            mapping_notes.append(mapping_note)
            report["layout"] = "hierarchical_expense"
        if not lines and n_rows > 0:
            report["_empty_parse_rows"] = int(report.get("_empty_parse_rows", 0)) + n_rows
        if lines:
            sheet_label = sc.sheet_name if sc else file_path.stem
            all_lines.extend(lines)
            ingested_names.append(sheet_label)
            report["sheets_ingested"].append(
                {"sheet": sheet_label, "rows": len(lines), "strategy": "standard"}
            )

    if suffix in (".xlsx", ".xls"):
        report["sheets_skipped"] = _build_ingestion_skip_list(file_path, ingested_names)
    note = mapping_notes[0] if mapping_notes else None
    return all_lines, _enrich_ingestion_report(report, all_lines, note)


def parse_spend_file(
    file_path: Path,
    taxonomy: Dict,
    default_amount_type: str = "actual",
    reporting_currency: str = "INR",
    workbook_manifest: Dict[str, Any] | None = None,
    source_system_id: Optional[str] = None,
) -> List[NormalizedSpendLine]:
    lines, _report = parse_spend_file_with_report(
        file_path,
        taxonomy,
        default_amount_type=default_amount_type,
        reporting_currency=reporting_currency,
        workbook_manifest=workbook_manifest,
        source_system_id=source_system_id,
    )
    return lines


def _load_json_records(file_path: Path) -> List[Dict[str, Any]]:
    raw = json.loads(file_path.read_text(encoding="utf-8"))
    if isinstance(raw, list):
        return [x for x in raw if isinstance(x, dict)]
    if isinstance(raw, dict):
        for key in ("data", "records", "rows", "lines", "spend", "transactions"):
            val = raw.get(key)
            if isinstance(val, list):
                return [x for x in val if isinstance(x, dict)]
        if all(isinstance(v, dict) for v in raw.values()):
            return list(raw.values())
    raise ValueError(f"{file_path.name}: JSON must be an array of objects or contain a data/records array")


def parse_spend_json_with_report(
    file_path: Path,
    taxonomy: Dict,
    default_amount_type: str = "actual",
    reporting_currency: str = "INR",
    source_system_id: Optional[str] = None,
) -> Tuple[List[NormalizedSpendLine], Dict[str, Any]]:
    records = _load_json_records(file_path)
    if not records:
        raise ValueError(f"{file_path.name}: no spend records found in JSON")
    frame = pd.DataFrame(records)
    lines, note = _dataframe_to_spend_lines(
        frame,
        file_path,
        taxonomy,
        default_amount_type=default_amount_type,
        reporting_currency=reporting_currency,
        source_system_id=source_system_id,
    )
    report: Dict[str, Any] = {
        "source_file": file_path.name,
        "format": "json",
        "rows_parsed": len(lines),
        "column_mapping_note": note,
    }
    return lines, report


def parse_spend_json(
    file_path: Path,
    taxonomy: Dict,
    default_amount_type: str = "actual",
    reporting_currency: str = "INR",
    source_system_id: Optional[str] = None,
) -> List[NormalizedSpendLine]:
    lines, _ = parse_spend_json_with_report(
        file_path,
        taxonomy,
        default_amount_type=default_amount_type,
        reporting_currency=reporting_currency,
        source_system_id=source_system_id,
    )
    return lines


def _ocr_image(file_path: Path, lang: str = "eng") -> str:
    """OCR a single image file. Returns extracted text."""
    try:
        img = _PILImage.open(file_path)
        return _pytesseract.image_to_string(img, lang=lang) or ""
    except Exception as exc:
        logger.warning("OCR failed for %s: %s", file_path.name, exc)
        return ""


def _ocr_pdf(file_path: Path, lang: str = "eng") -> str:
    """OCR an image-only PDF by converting pages to images then running Tesseract.

    Requires pdf2image (pip install pdf2image) and poppler system library.
    Falls back to empty string if pdf2image is not available.
    """
    try:
        from pdf2image import convert_from_path  # type: ignore[import]
    except ImportError:
        logger.warning("pdf2image not installed — cannot OCR image-only PDF %s", file_path.name)
        return ""
    try:
        images = convert_from_path(str(file_path), dpi=200)
        parts = [_pytesseract.image_to_string(img, lang=lang) or "" for img in images]
        return "\n".join(parts)
    except Exception as exc:
        logger.warning("PDF OCR failed for %s: %s", file_path.name, exc)
        return ""


def transliterate_to_latin(text: str, source_script: str = "DEVANAGARI") -> str:
    """Transliterate Indian-script text to Latin (ITRANS scheme) for script-agnostic matching.

    Returns original text unchanged when indic-transliteration is not installed.
    """
    if not _TRANSLITERATION_AVAILABLE or not text:
        return text
    try:
        script_map = {
            "DEVANAGARI": _sanscript.DEVANAGARI,
            "TAMIL": _sanscript.TAMIL,
            "TELUGU": _sanscript.TELUGU,
            "KANNADA": _sanscript.KANNADA,
            "BENGALI": _sanscript.BENGALI,
            "GUJARATI": _sanscript.GUJARATI,
            "GURMUKHI": _sanscript.GURMUKHI,
        }
        src = script_map.get(source_script.upper(), _sanscript.DEVANAGARI)
        return _transliterate(text, src, _sanscript.ITRANS)
    except Exception:
        return text


def parse_document(file_path: Path, ocr_lang: str = "eng") -> str:
    """Parse a document file and return its text content.

    ocr_lang — Tesseract language code(s) for image-based PDFs and images.
    Supports multi-language strings, e.g. 'eng+hin' for Hindi invoices.
    Common codes: hin, tam, tel, kan, mar, ben, guj, pan.
    Falls back to text extraction (no OCR) when pytesseract is not installed.
    """
    suffix = file_path.suffix.lower()
    if suffix == ".txt":
        return file_path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".docx":
        doc = Document(file_path)
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    if suffix == ".pdf":
        reader = PdfReader(str(file_path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        text = "\n".join(pages).strip()
        # If pypdf extracted <20 chars (image-only PDF), try OCR fallback.
        if len(text) < 20 and _OCR_AVAILABLE:
            text = _ocr_pdf(file_path, lang=ocr_lang)
        return text
    if suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp"):
        if _OCR_AVAILABLE:
            return _ocr_image(file_path, lang=ocr_lang)
        logger.warning("OCR requested for image file but pytesseract is not installed: %s", file_path.name)
        return ""
    if suffix in (".csv", ".xlsx", ".xls"):
        if suffix == ".csv":
            frame = pd.read_csv(file_path)
        else:
            frame = pd.read_excel(file_path)
        return frame.head(10).to_csv(index=False)
    return ""
