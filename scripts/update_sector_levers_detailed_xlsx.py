#!/usr/bin/env python3
"""Regenerate Sector_Levers_Detailed.xlsx from skills/sector-packs/*/sector_levers.json."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SECTOR_DIR = ROOT / "skills" / "sector-packs"
OUT_PATH = ROOT / "Sector_Levers_Detailed.xlsx"

# Sheet name (Excel 31-char limit) -> pack_id
SHEET_PACK_MAP: List[Tuple[str, str]] = [
    ("01. bfsi_banks", "bfsi_banks"),
    ("02. conglomerate", "conglomerate"),
    ("03. energy_utilities", "energy_utilities"),
    ("04. financial_services_n", "financial_services_nonbank"),
    ("05. fmcg_consumer", "fmcg_consumer"),
    ("06. gcc_capability_cente", "gcc_capability_centers"),
    ("07. healthcare_hospitals", "healthcare_hospitals"),
    ("08. hospitality_travel", "hospitality_travel"),
    ("09. insurance_general", "insurance_general"),
    ("10. it_ites", "it_ites"),
    ("11. manufacturing_divers", "manufacturing_diversified"),
    ("12. pharma_lifesciences", "pharma_lifesciences"),
    ("13. psu_cpse", "psu_cpse"),
    ("14. retail_organized", "retail_organized"),
    ("15. telecom_infra", "telecom_infra"),
]

HEADERS = [
    "Lever ID",
    "Lever Name",
    "Family",
    "Description",
    "p10 %",
    "p50 %",
    "p90 %",
    "Impl. Weeks",
    "Type",
    "Complexity",
    "Exec. Prob.",
    "Sustain. Score",
    "Key Conditions",
    "Bounce-Back Risk",
    "Industry Sources",
    "Applicable If",
    "Execution Playbook",
    "Diagnostic Signals",
    "Required Data Fields",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")


def _fmt_list(items: List[str]) -> str:
    return "; ".join(str(x) for x in items if x)


def _fmt_playbook(steps: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for i, step in enumerate(steps or [], 1):
        owner = step.get("owner_role", "")
        weeks = step.get("duration_weeks", "")
        deps = step.get("dependencies")
        dep_note = f" [after: {deps[0]}]" if deps else ""
        lines.append(f"{i}. ({owner}, {weeks}w{dep_note}) {step.get('step', '')}")
    return "\n".join(lines)


def _fmt_diagnostics(signals: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for sig in signals or []:
        lines.append(
            f"• {sig.get('signal', '')} [{sig.get('evidence_source', '')}] → {sig.get('confirms', '')}"
        )
    return "\n".join(lines)


def _lever_row(lever: Dict[str, Any]) -> List[Any]:
    sr = lever.get("savings_range_pct") or {}
    impl = lever.get("implementation_weeks") or {}
    conditions = lever.get("condition_precedents") or []
    sources = lever.get("industry_sources") or []
    applicable = lever.get("applicable_if") or []
    desc = conditions[0] if conditions else None
    return [
        lever.get("lever_id", ""),
        lever.get("lever_name", ""),
        lever.get("lever_family", ""),
        desc,
        sr.get("p10"),
        sr.get("p50"),
        sr.get("p90"),
        impl.get("p50"),
        lever.get("savings_type", ""),
        lever.get("complexity_tier", ""),
        lever.get("base_execution_probability"),
        lever.get("sustainability_score"),
        _fmt_list(conditions),
        lever.get("bounce_back_risk", ""),
        _fmt_list(sources),
        _fmt_list(applicable),
        _fmt_playbook(lever.get("execution_playbook") or []),
        _fmt_diagnostics(lever.get("diagnostic_signals") or []),
        _fmt_list(lever.get("required_data_fields") or []),
    ]


def _style_sheet(ws, ncols: int, data_rows: int) -> None:
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 32
    for col in range(1, ncols + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = {
        1: 28, 2: 36, 3: 12, 4: 28, 5: 8, 6: 8, 7: 8, 8: 10,
        9: 12, 10: 12, 11: 10, 12: 12, 13: 40, 14: 14, 15: 50,
        16: 36, 17: 55, 18: 45, 19: 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(4, 4 + data_rows):
        ws.row_dimensions[row].height = 72
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).alignment = WRAP
    ws.freeze_panes = "A4"


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, pack_id in SHEET_PACK_MAP:
        path = SECTOR_DIR / pack_id / "sector_levers.json"
        data = json.loads(path.read_text(encoding="utf-8"))
        levers = data.get("sector_specific_levers", [])
        sector_title = data.get("sector", pack_id.replace("_", " ").title())

        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value=f"{sector_title} — Detailed Savings Levers")
        ws.cell(row=1, column=1).font = TITLE_FONT
        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=3, column=col, value=header)
        for i, lever in enumerate(levers, 4):
            for col, val in enumerate(_lever_row(lever), 1):
                ws.cell(row=i, column=col, value=val)
        _style_sheet(ws, len(HEADERS), len(levers))

    return wb


def main() -> None:
    if (ROOT / "~$Sector_Levers_Detailed.xlsx").exists():
        print(
            "WARNING: Excel appears to have Sector_Levers_Detailed.xlsx open "
            "(~$ lock file present). Close Excel if save fails."
        )
    wb = build_workbook()
    wb.save(OUT_PATH)
    total = sum(
        len(json.loads((SECTOR_DIR / pid / "sector_levers.json").read_text())["sector_specific_levers"])
        for _, pid in SHEET_PACK_MAP
    )
    print(f"Wrote {OUT_PATH} — {len(SHEET_PACK_MAP)} sheets, {total} sector levers")


if __name__ == "__main__":
    main()
